import tkinter as tk
from openpyxl import Workbook


class ExcelFile:

    def __init__(self):
        self.wb = Workbook()
        self.window = tk.Tk()
        self.window.title("New workbook")
        self.ws = ""
        self.month_title = ""
        self.date_c_header = ""
        self.location_c_header = ""
        self.amount_c_header = ""

    def get_ws_title(self):
        return self.month_title

    def name_worksheet(self):
        # create a new workbook titled with the month
        tk.Label(self.window, text = "Month to Log").grid(row=0)
        self.entry = tk.Entry(self.window)
        self.entry.grid(row=1)
        self.enter_btn = tk.Button(self.window, text = "Enter", command = self.handle_enter)
        self.enter_btn.grid(row=2)

    def handle_enter(self, event = None):
        # sets the title as a class attribute
        # helper function to name_worksheet()
        self.month_title = self.entry.get()  # self.month_title is the title of the worksheet

    def create_new_ws(self):
        # creates a new worksheet in the workbook with self.month_title as the Worksheet title
        self.ws = self.wb.active
        self.ws = self.wb.create_sheet(self.month_title)

    def change_title(self, title):
        # changes the worksheet's title
        self.month_title = title    # updates the class attribute
        self.ws.title = title       # changes the name of the sheet

    def initialise_ws(self):
        # sets up the workbook headers
        self.date_c_header = self.ws['A1']
        self.date_c_header.value = "DATE"
        self.location_c_header = self.ws['B1']
        self.location_c_header.value = "LOCATION"
        self.amount_c_header = self.ws['C1']
        self.amount_c_header.value = "AMOUNT"

    def set_wbfile_name(self):
        # prompt user to get a file name
        self.wbfilename = input("Enter the workbook file name: ")
        return self.wbfilename

    def save_wb(self, title):
        self.file_name = title + ".xlsx"
        self.wb.save(self.file_name)

