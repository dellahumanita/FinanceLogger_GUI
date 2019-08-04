from openpyxl import Workbook
from openpyxl import load_workbook

# NOTE:
# - function name uses full form, i.e. "workbook", "worksheet", etc.
# - variables and class attributes uses shorted form, i.e. "wb", "ws", etc.


class ExcelFile:
    def __init__(self):
        self.title = ""
        self.new_wb = None  # new workbook
        self.ws = None  # new worksheet
        self.load_wb = None  # existing workbook

    def new_workbook(self):
        # create a new worksheet
        self.new_wb = Workbook()
        self.ws = self.new_wb.active
        print("new wb created")

    def workbook_title(self, title):
        # sets the month to log as title of workbook
        self.title = title
        print(self.title, "is set title...")

    def open_workbook(self, filename):
        # open an existing workbook
        self.load_wb = load_workbook(filename)
        self.ws = self.load_wb.active
        print("Opening workbook...")

    def initialise_worksheet(self):
        # setting up workbook headers

        date_c_header = self.ws['A1']
        location_c_header = self.ws['B1']
        amount_c_header = self.ws['C1']

        date_c_header.value = "DATE"
        location_c_header.value = "LOCATION"
        amount_c_header.value = "AMOUNT"

    def add_log(self, date, location, amount):
        # adds log to the following row

        last_row = self.find_last_row_used()
        next_row = last_row+1
        self.ws.cell(column=1, row=next_row, value=date)
        self.ws.cell(column=2, row=next_row, value=location)
        self.ws.cell(column=3, row=next_row, value=amount)
        print("Added log...")

    def find_last_row_used(self):
        # searches for the last row entered
        return self.ws.max_row

    def save_new_workbook(self):
        # saves workbook
        file_name = self.title + ".xlsx"
        self.new_wb.save(file_name)

    def save_opened_workbook(self):
        file_name = self.title + ".xlsx"
        self.load_wb.save(file_name)
        print(file_name, "is the saved file.")

    def view(self):  #todo
        # iterate over used cells in Excel file
        # add to data structure -- ?
        # convert to string
        # print in separate window

        self.ws = self.load_wb.get_sheet_by_name("Sheet")
        print(self.ws)
        file_data = {}
        print("Reading rows...")
        for row in range(2, self.ws.max_row + 1):
            date = self.ws['A' + str(row)].value
            location = self.ws['B' + str(row)].value
            amount = self.ws['C' + str(row)].value

            file_data.setdefault(date, {location, amount})
            print(file_data)

# xl_new = ExcelFile()
# xl_new.new_workbook()
# xl_new.workbook_title("testfile")
# xl_new.initialise_worksheet()
# xl_new.add_log('01/02/2019','No Frills','30')
# xl_new.add_log('02/03/2019','no','10')
# xl_new.add_log('03/04/2019', 'Chapters', '100')
#
# xl_new.save_workbook()


# xl = ExcelFile()
# xl.open_workbook('May.xlsx')
# xl.workbook_title('May')
# xl.add_log("02/02/2019", "Shoppers", "15")
# xl.save_opened_workbook()
